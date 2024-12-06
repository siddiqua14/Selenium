import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from requests.exceptions import RequestException

# Function to create or load the Excel workbook and sheet
def create_excel_report(test_case):
    try:
        # Try to load the existing workbook
        workbook = openpyxl.load_workbook("test_results.xlsx")
    except (FileNotFoundError, KeyError):
        # If the file is missing or corrupted, create a new workbook
        workbook = openpyxl.Workbook()

    # Create the "url_status" sheet or load it if it already exists
    if "url_status" not in workbook.sheetnames:
        sheet = workbook.create_sheet("url_status")
        # Write headers
        sheet.append(["Page URL", "Test Case", "Result", "Comments"])
    else:
        sheet = workbook["url_status"]
        # Clear all rows except headers
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

    return workbook, sheet

# Function to write results into the Excel sheet
def write_report(test_case, result, comments, page_url):
    workbook, sheet = create_excel_report(test_case)
    sheet.append([page_url, test_case, result, comments])
    workbook.save("test_results.xlsx")

# Initialize WebDriver
options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Function to check the status code of a URL
def check_url_status(url):
    try:
        response = requests.head(url, allow_redirects=True, timeout=10)
        if response.status_code == 404:
            return "Fail", f"Status Code: {response.status_code} (Not Found)"
        elif response.status_code == 200:
            return "Pass", f"Status Code: {response.status_code} (OK)"
        else:
            return "Warn", f"Status Code: {response.status_code} (Unexpected)"
    except RequestException as e:
        return "Fail", f"Error: {e}"

# Main function to check links and run tests
def check_links_and_run_tests():
    # Open the main page
    driver.get("https://www.alojamiento.io/")
    
    # Wait for the body element to be present on the page
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

    # Collect all the links on the page
    links = driver.find_elements(By.TAG_NAME, 'a')
    link_urls = [link.get_attribute('href') for link in links if link.get_attribute('href')]

    # Check the status code of all connected URLs
    for url in link_urls:
        if url:  # Ensure it's not an empty URL
            print(f"Checking URL: {url}")
            result, comments = check_url_status(url)
            write_report("URL Status Code test", result, comments, url)
            print(f"Test result for {url}: {result} - {comments}")

    # Close the browser after tests
    try:
        driver.quit()
    except WebDriverException as e:
        print(f"Error closing the browser: {e}")

# Run the test
check_links_and_run_tests()
