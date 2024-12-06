import time
import openpyxl
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, ElementNotInteractableException

# Set up logging for real-time feedback
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")

# Function to create or load the Excel workbook and sheet
def create_excel_report(test_case):
    try:
        workbook = openpyxl.load_workbook("test_results.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Create the "currency_filtering" sheet or load it if it already exists
    if "currency_filtering" not in workbook.sheetnames:
        sheet = workbook.create_sheet(title="currency_filtering")
        # Write headers for the sheet
        sheet.append(["Test Case", "Currency", "Result", "Comments"])
    else:
        sheet = workbook["currency_filtering"]
    
    return workbook, sheet

# Function to write results into the Excel file
def write_report(test_case, currency, result, comments):
    workbook, sheet = create_excel_report(test_case)
    sheet.append([test_case, currency, result, comments])
    workbook.save("test_results.xlsx")

# Initialize WebDriver
options = Options()
options.add_argument("--headless")  # Run in headless mode for automation
options.add_argument("--disable-gpu")  # Disable GPU for smoother operation
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Function to get all available currencies from the dropdown
def get_available_currencies():
    try:
        # Open the currency dropdown in the footer
        currency_filter = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "footer-currency-dd"))
        )
        driver.execute_script("arguments[0].click();", currency_filter)  # Click to open the dropdown

        # Find all currency options
        currency_options = driver.find_elements(By.XPATH, "//ul[@class='select-ul']/li")
        
        # Extract the currency values from the options
        currencies = [option.text.strip() for option in currency_options]
        logging.info(f"Available currencies: {currencies}")
        return currencies
    except Exception as e:
        logging.error(f"Error getting currencies: {e}")
        return []

# Function to perform currency filtering (click dropdown and select currency)
def perform_currency_filtering(currency):
    try:
        # Locate the currency dropdown in the footer
        currency_filter = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "footer-currency-dd"))
        )

        # Use JavaScript to click if Selenium's regular click doesn't work
        driver.execute_script("arguments[0].click();", currency_filter)

        # Wait until the options are visible
        currency_options = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//ul[@class='select-ul']/li"))
        )

        # Find and click the currency option
        for option in currency_options:
            currency_text = option.text.strip()
            if currency in currency_text:
                driver.execute_script("arguments[0].click();", option)  # Use JavaScript to click the desired currency
                logging.info(f"Currency changed to: {currency}")
                time.sleep(3)  # Wait for the page to reload with the new currency
                return True
        logging.warning(f"Currency option {currency} not found.")
        return False
    except ElementNotInteractableException as e:
        logging.error(f"Error interacting with the dropdown: {e}")
        return False
    except Exception as e:
        logging.error(f"Error in currency filtering: {e}")
        return False

# Function to check property tiles' currency after filtering
def check_property_tiles_currency(currency):
    # Check that the currency displayed in the property tiles matches the selected currency
    property_tiles = driver.find_elements(By.CLASS_NAME, "property-tile")  # Adjust this to the correct class name

    for tile in property_tiles:
        try:
            currency_text = tile.find_element(By.CLASS_NAME, "currency").text  # Adjust the class name accordingly
            if currency_text != currency:
                write_report("Currency Filtering Test", currency, "Fail", f"Currency mismatch in property tile: {currency_text}")
                logging.error(f"Currency mismatch in property tile: {currency_text}")
                return False
        except Exception as e:
            write_report("Currency Filtering Test", currency, "Fail", "Error retrieving currency from property tile.")
            logging.error("Error retrieving currency from property tile.")
            return False
    
    # If all property tiles show the correct currency
    write_report("Currency Filtering Test", currency, "Pass", f"Currency successfully changed to {currency}")
    logging.info(f"Currency successfully changed to {currency}")
    return True

# Main function to run the test on the specific link
def run_currency_filtering_tests():
    # Define the link to test (only one link)
    target_url = "https://www.alojamiento.io/"
    
    # Visit the specific link
    logging.info(f"Visiting the target URL: {target_url}")
    driver.get(target_url)
    time.sleep(3)  # Wait for the page to load completely

    # Get all available currencies
    currencies_to_test = get_available_currencies()

    if not currencies_to_test:
        logging.error("No currencies found in the dropdown.")
        return

    # Loop through each currency and perform the test
    for currency in currencies_to_test:
        logging.info(f"Testing currency: {currency}")
        # Step 1: Perform currency filtering
        if perform_currency_filtering(currency):
            # Step 2: Check property tiles for currency change
            check_property_tiles_currency(currency)
        else:
            write_report("Currency Filtering Test", currency, "Fail", f"Failed to change currency to {currency}")
            logging.error(f"Failed to change currency to {currency}")

# Run the currency filtering tests for the specific link
run_currency_filtering_tests()

# Close the browser
try:
    driver.quit()
    logging.info("Browser closed successfully.")
except WebDriverException as e:
    logging.error(f"Error closing the browser: {e}")
