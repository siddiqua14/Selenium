import os
import openpyxl

def write_report(test_case, page_url, results, headers=None):
    """
    Write test results to an Excel file, with support for multiple test types.
    """
    file_name = "data/test_results.xlsx"
    directory = os.path.dirname(file_name)

    # Create the directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Load or create the workbook
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()

    # If the sheet for the given test_case already exists, delete it
    if test_case in workbook.sheetnames:
        del workbook[test_case]

    # Create a new sheet for the test case
    sheet = workbook.create_sheet(title=test_case)

    # Define default headers based on the type of test
    if test_case == "Currency filter test":
        if not headers:
            headers = ["Currency", "Default Price", "Tile Prices", "Status"]
    else:
        # For other tests (H1 tag, HTML validation, etc.), use a default header
        if not headers:
            headers = ["Page URL", "Test Case", "Result", "Comments"]

    # Add headers to the sheet
    sheet.append(headers)

    # Add rows to the sheet
    for result in results:
        row = [result.get(header, "N/A") for header in headers]
        sheet.append(row)

    # Save the workbook
    workbook.save(file_name)
    print(f"Results saved to {file_name}.")
