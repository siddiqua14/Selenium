import os
import openpyxl

def write_report(site_url, site_name, campaign_id, browser, country_code, ip, result):
    # Define the file name
    file_name = "data/scraped_data_report.xlsx"
    
    # Create or load an existing workbook
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    
    # Get or create a sheet
    if 'Scrape Data' not in wb.sheetnames:
        ws = wb.create_sheet('Scrape Data')
    else:
        ws = wb['Scrape Data']

    # Clear existing data (except headers)
    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

    # Write headers
    headers = ['Site URL', 'Site Name', 'Campaign ID', 'Browser', 'Country Code', 'IP Address', 'Result']
    ws.delete_rows(1, ws.max_row)  # Delete all rows before rewriting
    ws.append(headers)

    # Overwrite the data row
    ws.append([site_url, site_name, campaign_id, browser, country_code, ip, result])

    # Save the workbook
    wb.save(file_name)
