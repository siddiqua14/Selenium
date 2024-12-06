import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import WebDriverException

# Function to create or load the Excel workbook and sheet
def create_excel_report(test_case):
    try:
        workbook = openpyxl.load_workbook("test_results.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    if test_case not in workbook.sheetnames:
        sheet = workbook.create_sheet(test_case)
        sheet.append(["Page URL", "Test Case", "Result", "Comments"])
    else:
        sheet = workbook[test_case]
    
    return workbook, sheet

# Function to write results into the Excel sheet
def write_report(test_case, result, comments, page_url):
    workbook, sheet = create_excel_report(test_case)
    sheet.append([page_url, test_case, result, comments])
    workbook.save("test_results.xlsx")

# Initialize WebDriver
options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Function to perform HTML tag sequence test on a page
def run_html_tag_sequence_test_on_page(page_url):
    driver.get(page_url)
    time.sleep(5)  # Wait for the page to load

    tags = ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']
    missing_tags = []
    broken_sequence = False

    last_tag_index = -1
    for tag in tags:
        try:
            driver.find_element(By.TAG_NAME, tag)
            tag_index = tags.index(tag)
            if tag_index < last_tag_index:
                broken_sequence = True
            last_tag_index = tag_index
        except:
            missing_tags.append(tag)

    if missing_tags or broken_sequence:
        result = "Fail"
        comments = f"Missing tags: {', '.join(missing_tags)}"
        if broken_sequence:
            comments += " | Sequence broken."
    else:
        result = "Pass"
        comments = "All tags from H1 to H6 found in sequence."

    write_report("HTML tag sequence test", result, comments, page_url)

# Main function to check unique links and run tests
def check_links_and_run_tests():
    driver.get("https://www.alojamiento.io/")
    time.sleep(5)

    links = driver.find_elements(By.TAG_NAME, 'a')
    link_urls = {link.get_attribute('href') for link in links if link.get_attribute('href')}

    # Skip social media URLs
    excluded_domains = ["facebook.com", "twitter.com", "linkedin.com", "instagram.com"]
    filtered_links = [url for url in link_urls if not any(domain in url for domain in excluded_domains)]

    for link in filtered_links:
        run_html_tag_sequence_test_on_page(link)

    try:
        driver.quit()
    except WebDriverException as e:
        print(f"Error closing the browser: {e}")

# Run the test
check_links_and_run_tests()
