import time
import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import WebDriverException

# Function to create or load the Excel workbook and sheet
def create_excel_report(test_case):
    try:
        workbook = openpyxl.load_workbook("test_results.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Create the "url_status" sheet or load it if it already exists
    if "url_status" not in workbook.sheetnames:
        sheet = workbook.create_sheet(title="url_status")
        # Write headers for the sheet
        sheet.append(["Page URL", "Test Case", "Result", "Comments"])
    else:
        sheet = workbook["url_status"]
    
    return workbook, sheet

# Function to write results into the Excel file
def write_report(test_case, result, comments, page_url):
    workbook, sheet = create_excel_report(test_case)
    sheet.append([page_url, test_case, result, comments])
    workbook.save("test_results.xlsx")

# Initialize WebDriver
options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Function to check the status code of a URL
def check_url_status(url):
    try:
        response = requests.head(url, allow_redirects=True, timeout=10)
        if response.status_code == 404:
            return "Fail", f"Status Code: 404 (Not Found)"
        elif response.status_code == 200:
            return "Pass", f"Status Code: 200 (OK)"
        else:
            return "Warn", f"Status Code: {response.status_code} (Unexpected)"
    except requests.exceptions.RequestException as e:
        return "Fail", f"Error: {e}"

# Function to check all links and run the URL status code test
def check_links_and_run_tests():
    driver.get("https://www.alojamiento.io/")
    page_url = driver.current_url

    # Find all links on the page
    links = driver.find_elements(By.TAG_NAME, "a")
    valid_links = [link.get_attribute('href') for link in links if link.get_attribute('href') is not None]

    # Use a set to ensure that we check each URL only once
    unique_links = set(valid_links)

    # Run the URL status code test for each unique link
    for link in unique_links:
        run_url_status_test_on_page(link)

# Function to check URL status for a given page
def run_url_status_test_on_page(page_url):
    test_case = "URL Status Code Test"
    result, comments = check_url_status(page_url)
    write_report(test_case, result, comments, page_url)
    print(f"Test result for {page_url}: {result} - {comments}")

# Run the tests on all unique links
check_links_and_run_tests()

# Close the browser
try:
    driver.quit()
except WebDriverException as e:
    print(f"Error closing the browser: {e}")
