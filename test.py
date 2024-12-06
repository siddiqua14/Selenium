import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import WebDriverException, StaleElementReferenceException

# Function to create or load the Excel workbook and sheet
def create_excel_report(test_case):
    try:
        workbook = openpyxl.load_workbook("test_results.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Create a new sheet for the test if it doesn't exist
    if test_case not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=test_case)
        # Write headers for the sheet
        sheet.append(["Page URL", "Test Case", "Result", "Comments"])
    else:
        sheet = workbook[test_case]
    
    return workbook, sheet

# Function to write results into the Excel file
def write_report(test_case, result, comments, page_url):
    workbook, sheet = create_excel_report(test_case)
    sheet.append([page_url, test_case, result, comments])
    workbook.save("test_results.xlsx")

# Initialize WebDriver
options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Function to check all links and run the H1 tag existence test
def check_links_and_run_tests():
    driver.get("https://www.alojamiento.io/")
    page_url = driver.current_url

    # Find all links on the page
    links = driver.find_elements(By.TAG_NAME, "a")
    valid_links = [link.get_attribute('href') for link in links if link.get_attribute('href') is not None]

    # Run the H1 tag existence test for each link
    for link in valid_links:
        run_h1_tag_test_on_page(link)

# Function to check H1 tag existence for a given page
def run_h1_tag_test_on_page(page_url):
    driver.get(page_url)
    time.sleep(2)  # Wait for the page to load completely

    # Test 1: H1 tag existence
    test_case = "H1 tag existence test"
    try:
        driver.find_element(By.TAG_NAME, 'h1')
        write_report(test_case, "Pass", "H1 tag found.", page_url)
        print(f"Test passed: H1 tag exists on {page_url}.")
    except:
        write_report(test_case, "Fail", "H1 tag missing.", page_url)
        print(f"Test failed: H1 tag missing on {page_url}.")

# Run the tests on all links
check_links_and_run_tests()

# Close the browser
try:
    driver.quit()
except WebDriverException as e:
    print(f"Error closing the browser: {e}")
