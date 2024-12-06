import time
import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import WebDriverException

# Function to create or load the Excel workbook and sheet
def create_excel_report():
    try:
        workbook = openpyxl.load_workbook("test_results.xlsx")
        if "Test Cases H1-H6-Image" not in workbook.sheetnames:
            sheet = workbook.create_sheet(title="Test Cases H1-H6-Image")  # Create a new sheet if it doesn't exist
        else:
            sheet = workbook["Test Cases H1-H6-Image"]
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Write headers for the first sheet
        sheet.title = "Test Cases H1-H6-Image"  # Name the sheet
        sheet.append(["Page URL", "Test Case", "Result", "Comments"])
    
    return workbook, sheet

# Function to write results into the Excel file
def write_report(test_case, result, comments, page_url):
    workbook, sheet = create_excel_report()
    sheet.append([page_url, test_case, result, comments])
    workbook.save("test_results.xlsx")

# Initialize WebDriver
options = Options()
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Open vacation rental website
driver.get("https://www.alojamiento.io/")
page_url = driver.current_url

# Test: H1 tag existence
test_case = "H1 tag existence test"
try:
    driver.find_element(By.TAG_NAME, 'h1')
    write_report(test_case, "Pass", "H1 tag found.", page_url)
    print("Test passed: H1 tag exists.")
except:
    write_report(test_case, "Fail", "H1 tag missing.", page_url)
    print("Test failed: H1 tag missing.")

# Test: HTML tag sequence [H1-H6]
test_case = "HTML tag sequence test"
tags = ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']
missing_tags = []
for tag in tags:
    try:
        driver.find_element(By.TAG_NAME, tag)
    except:
        missing_tags.append(tag)

if missing_tags:
    write_report(test_case, "Fail", f"Missing tags: {', '.join(missing_tags)}", page_url)
    print(f"Test failed: Missing tags - {', '.join(missing_tags)}")
else:
    write_report(test_case, "Pass", "All tags from H1 to H6 found.", page_url)
    print("Test passed: All tags H1 to H6 are present.")

# Test: Image alt attribute check
test_case = "Image alt attribute test"
images = driver.find_elements(By.TAG_NAME, 'img')
missing_alt = [img.get_attribute('src') for img in images if not img.get_attribute('alt')]
if missing_alt:
    write_report(test_case, "Fail", f"Images missing alt attribute: {', '.join(missing_alt)}", page_url)
    print(f"Test failed: Missing alt attribute for images - {', '.join(missing_alt)}")
else:
    write_report(test_case, "Pass", "All images have alt attributes.", page_url)
    print("Test passed: All images have alt attributes.")

# Close the browser
try:
    driver.quit()
except WebDriverException as e:
    print(f"Error closing the browser: {e}")
